# from flask import Flask
# from redis import Redis

# app = Flask(__name__)
# redis = Redis(host='redis', port=6379)

# @app.route('/')
# def hello():
#     redis.incr('hits')
#     counter = str(redis.get('hits'),'utf-8')
#     return "This webpage has been viewed "+counter+" time(s)"

# if __name__ == "__main__":
#     app.run(host="0.0.0.0", port=8000, debug=True)



############################################# 
#### Example 2: with name in the uri ########
############################################# 

## The uri for this to run will be : http://localhost:8000/?name=Sofia
## Github repository: https://github.com/docker/awesome-compose/tree/master/flask-redis

# from flask import Flask, request
# from datetime import datetime

# app = Flask(__name__)

# @app.route('/')
# def hello():
#     name = request.args['name']
#     return HELLO_HTML.format(
#              name, str(datetime.now()))

# HELLO_HTML = """
#      <html><body>
#          <h1>Hello, {0}!</h1>
#          The time is {1}.
#      </body></html>"""

# if __name__ == "__main__":
#     # Launch the Flask dev server
#     app.run(host="0.0.0.0", port=8000, debug=True)


############################################ 
#### Example 3: buttons and actions ########
############################################ 

### Good link : https://plainenglish.io/blog/how-to-create-a-basic-form-in-python-flask-af966ee493fa
# templates should be under the template folder
### Templates : https://www.digitalocean.com/community/tutorials/how-to-use-templates-in-a-flask-application

from crypt import methods
from pyexpat.errors import messages
from flask import Flask, render_template, request
from flask import jsonify, redirect, url_for
from flask import send_file, make_response

# Libraries for DNAC start: 
#from asyncore import write
from csv import writer
import site
from turtle import shape
from urllib import response
from dnacentersdk import api
import json
import os
from openpyxl import Workbook
# import pandas as pd
# import numpy as np

# Libraries for DNAC end


app = Flask(__name__)

@app.route("/input", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        print(request.form["interface"])
        print(request.form["ip"])
        print(request.form["subnet"])
        return 'OK' 

    return render_template("home.html")

@app.route("/submit-form", methods=['POST'])
def submit_form():
    # get data from the form
    interface = request.form["interface"]
    ip = request.form["ip"]
    subnet = request.form["subnet"]

    wb = Workbook()
    ws = wb.active

    ws['A1'] = "Interface"
    ws['B1'] = 'IP'
    ws['C1'] = 'Subnet'
    ws['A2'] = interface
    ws['B2'] = ip
    ws['C2'] = subnet

    wb.save('test_data.xlsx')

    return render_template("download.html")

@app.route('/download')
def download_file():

    path = 'test_data.xlsx'
    response = make_response(send_file(path,as_attachment=True))
    response.headers['Content-Disposition'] = 'attachment; filename=test_data.xlsx'
    
    return response

@app.route('/')
def about():
    return render_template('about.html')

@app.route('/comments/')
def comments():
    comments = ['Choose 1.',
                'This is the second comment.',
                'This is the third comment.',
                'This is the fourth comment.'
                ]

    return render_template('comments.html', comments=comments)


# DNAC begin:

@app.route('/test-button')
def button():
    return render_template('button.html')

@app.route('/api/sites', methods=['GET'])
def get_sites():
    ##### Connection to DNAC ####
    dnac = api.DNACenterAPI(base_url='https://sandboxdnac.cisco.com', username='devnetuser', password='Cisco123!', verify=False)

    sites = dnac.sites.get_site()
    names= []
    for i in range(len(sites.response)):
        names.append(sites.response[i].name)
    return render_template('sites.html', response=names)
 

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=8007 )

# http://127.0.0.1:8000/comments/


# @app.route("/", methods=['GET', 'POST'])
# def index():
#     if request.method == 'POST':
#         if request.form.get('action1') == 'VALUE1':
#             pass # do something
#         elif  request.form.get('action2') == 'VALUE2':
#             pass # do something else
#         else:
#             pass # unknown
#     elif request.method == 'GET':
#         return render_template('index.html', form= form)
    
#     return render_template("index.html")